import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from urllib.parse import quote_plus
from sqlalchemy import create_engine
from datetime import datetime, timedelta
def count_images_today(directory):
    today = (datetime.today() - timedelta(days=1)).date()
    image_count = 0
    files_counted = set()
    extensions_image = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']
    if isinstance(directory, list):
        directories = directory
    else:
        directories = [directory]
    for dir_path in directories:
        for filename in os.listdir(dir_path):
            if any(filename.lower().endswith(ext) for ext in extensions_image):
                file_path = os.path.join(dir_path, filename)
                creation_time = datetime.fromtimestamp(os.path.getctime(file_path)).date()
                if creation_time == today:
                    prefix = filename[:5]
                    if prefix not in files_counted:
                        files_counted.add(prefix)
                        image_count += 1
    return image_count
directory_paths = {
    'Transcamino': 'C:\\Users\\lmsantos\\OneDrive - Ecourbis Ambiental SA\\Chamado 191667 - QR-Code Containeres\\01 - Transcamino\\FOTOS',
    'MC Lopes': 'C:\\Users\\lmsantos\\OneDrive - Ecourbis Ambiental SA\\Chamado 191667 - QR-Code Containeres\\02 - MC Lopes\\FOTOS',
    'Rodomarca': 'C:\\Users\\lmsantos\\OneDrive - Ecourbis Ambiental SA\\Chamado 191667 - QR-Code Containeres\\03 - Rodomarca\\FOTOS',
    'Dois Aranha': [
        'C:\\Users\\lmsantos\\OneDrive - Ecourbis Ambiental SA\\Chamado 191667 - QR-Code Containeres\\04 - Dois Aranha\\Fotos Containers Ecourbis QR Code - Leste',
        'C:\\Users\\lmsantos\\OneDrive - Ecourbis Ambiental SA\\Chamado 191667 - QR-Code Containeres\\04 - Dois Aranha\\Fotos Containers Ecourbis QR Code - Sul'
    ]
}
yesterday = (datetime.today() - timedelta(days=1)).date()
yesterday_str = yesterday.strftime('%d/%m/%Y')
results = {'DATA': yesterday_str}
conn_str = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    'SERVER=SEVSUL-22\\WEB;'
    'DATABASE=CONTEINERESSGC;'
    'Trusted_Connection=yes;'
)
conn_str = quote_plus(conn_str)
engine = create_engine(f'mssql+pyodbc:///?odbc_connect={conn_str}')
consulta_sql = """
SELECT NomeEmpresa, COUNT(*) AS Contagem
FROM tbConteineres
WHERE DataQrCode >= CAST(DATEADD(DAY, -1, GETDATE()) AS DATE) AND NumeroQrCode > '0'
GROUP BY NomeEmpresa
"""
results_df = pd.read_sql_query(consulta_sql, engine)
resultado_transcamino = 0
resultado_mc_lopes = 0
resultado_rodomarca = 0
resultado_dois_aranha = 0
resultado_ecourbis = 0
for index, row in results_df.iterrows():
    if row['NomeEmpresa'] == 'TRANSCAMINO':
        resultado_transcamino = row['Contagem']
    elif row['NomeEmpresa'] == 'MC LOPES':
        resultado_mc_lopes = row['Contagem']
    elif row['NomeEmpresa'] == 'RODOMARCA':
        resultado_rodomarca = row['Contagem']
    elif row['NomeEmpresa'] == 'DOIS ARANHA':
        resultado_dois_aranha = row['Contagem']
    elif row['NomeEmpresa'] == 'ECOURBIS AMBIENTAL SA':
        resultado_ecourbis = row['Contagem']
consulta_sql_cadastradas = """
SELECT COUNT(*)
FROM tbConteineres
WHERE DataQrCode >= CAST(DATEADD(DAY, -1, GETDATE()) AS DATE) AND NumeroQrCode > '0'
"""
resultado_cadastradas = pd.read_sql_query(consulta_sql_cadastradas, engine).iloc[0, 0]

consulta_sql_sul = """
SELECT COUNT(*)
FROM tbConteineres
WHERE DataQrCode >= CAST(DATEADD(DAY, -1, GETDATE()) AS DATE) AND NumeroQrCode > '0' and UnidadeId = '1'
"""
resultado_sul = pd.read_sql_query(consulta_sql_sul, engine).iloc[0, 0]

consulta_sql_leste = """
SELECT COUNT(*)
FROM tbConteineres
WHERE DataQrCode >= CAST(DATEADD(DAY, -1, GETDATE()) AS DATE) AND NumeroQrCode > '0' and UnidadeId = '2'
"""
resultado_leste = pd.read_sql_query(consulta_sql_leste, engine).iloc[0, 0]
for folder_name, directory_path in directory_paths.items():
    quantity_images = count_images_today(directory_path)
    results[folder_name] = quantity_images
results['Geral'] = sum(results[folder_name] for folder_name in directory_paths.keys())
results['Cadastradas_Ecourbis'] = resultado_ecourbis
results['Cadastradas_Transcamino'] = resultado_transcamino
results['Cadastradas_MC_Lopes'] = resultado_mc_lopes
results['Cadastradas_Rodomarca'] = resultado_rodomarca
results['Cadastradas_Dois_Aranha'] = resultado_dois_aranha
results['Cadastradas_Geral'] = resultado_cadastradas
results[''] = ''
results['Cadastrados_Sul'] = resultado_sul
results['Cadastrados_Leste'] = resultado_leste
df_results = pd.DataFrame([results])
excel_name = 'C:\\Users\\lmsantos\\Documents\\imagens_adicionadas.xlsx'
try:
    df_existent = pd.read_excel(excel_name, engine='openpyxl', header=1)
    df_existent = df_existent[df_existent['DATA'] != 'Total']
    df_final = pd.concat([df_existent, df_results], ignore_index=True)
except FileNotFoundError:
    df_final = df_results
total_line = {'DATA': 'Total'}
total_line['Transcamino'] = df_final['Transcamino'].sum()
total_line['MC Lopes'] = df_final['MC Lopes'].sum()
total_line['Rodomarca'] = df_final['Rodomarca'].sum()
total_line['Dois Aranha'] = df_final['Dois Aranha'].sum()
total_line['Cadastradas_Ecourbis'] = df_final['Cadastradas_Ecourbis'].sum()
total_line['Cadastradas_Transcamino'] = df_final['Cadastradas_Transcamino'].sum()
total_line['Cadastradas_MC_Lopes'] = df_final['Cadastradas_MC_Lopes'].sum()
total_line['Cadastradas_Rodomarca'] = df_final['Cadastradas_Rodomarca'].sum()
total_line['Geral'] = df_final['Geral'].sum()
total_line['Cadastradas_Geral'] = df_final['Cadastradas_Geral'].sum()
total_line[''] = ''
total_line['Cadastrados_Sul'] = df_final['Cadastrados_Sul'].sum()
total_line['Cadastrados_Leste'] = df_final['Cadastrados_Leste'].sum()
df_final.loc[len(df_final)] = total_line

wb = load_workbook(excel_name)
ws = wb["Sheet1"]
if 'Total' in df_final['DATA'].values:
    total_row_index = df_final[df_final['DATA'] == 'Total'].index[0] + 2
    for col_num in range(1, len(total_line) + 2):
        cell = ws.cell(row=total_row_index, column=col_num)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
for index, row in df_final.iterrows():
    for col_num, value in enumerate(row):
        cell = ws.cell(row=index + 3, column=col_num + 1)
        cell.value = value
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
if not any(['B1:F1' in str(range) for range in ws.merged_cells.ranges]):
    ws.merge_cells('B1:F1')
    blue_fill = PatternFill(start_color='90D5AC', end_color='90D5AC', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.fill = blue_fill
            cell.font = Font(bold=True)
    for cell in ws['A']:
        cell.font = Font(bold=True)
if not any(['G1:L1' in str(range) for range in ws.merged_cells.ranges]):
    ws.merge_cells('G1:L1')
    blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=7, max_col=12):
        for cell in row:
            cell.fill = blue_fill
            cell.font = Font(bold=True)
if not any(['N1:O1' in str(range) for range in ws.merged_cells.ranges]):
    ws.merge_cells('N1:O1')
    blue_fill = PatternFill(start_color='90D5AC', end_color='90D5AC', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=14, max_col=15):
        for cell in row:
            cell.fill = blue_fill
            cell.font = Font(bold=True)
total_row_index = len(df_final) + 3
for col_num in range(1, len(total_line) + 2):
    cell = ws.cell(row=total_row_index - 1, column=col_num)  # Linha total
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Aplica o amarelo
    cell.border = thin_border
    cell.font = Font(bold=True)

wb.save(excel_name)

wb = load_workbook(excel_name)
ws = wb["Sheet1"]
df = pd.read_excel(excel_name, engine='openpyxl', header=1)
df = df[df['DATA'] != 'Total']
columns_of_interest = ['Transcamino', 'MC Lopes', 'Rodomarca', 'Dois Aranha', 'Geral']
sums = df[columns_of_interest].sum()
columns_of_interest2 = ['Cadastradas_Ecourbis', 'Cadastradas_Transcamino', 'Cadastradas_MC_Lopes', 'Cadastradas_Rodomarca', 'Cadastradas_Dois_Aranha', 'Cadastradas_Geral']
sums2 = df[columns_of_interest2].sum()
if 'Resumo' not in wb.sheetnames:
    ws_resumo = wb.create_sheet('Resumo')
else:
    ws_resumo = wb['Resumo']
    for row in ws_resumo.iter_rows(min_row=1, max_row=ws_resumo.max_row, min_col=1, max_col=ws_resumo.max_column):
        for cell in row:
            if not any(cell.coordinate in m for m in ws_resumo.merged_cells.ranges):
                cell.value = None
# Seção 1: Fotos cadastradas
ws_resumo.merge_cells('A1:B1')
ws_resumo.cell(row=1, column=1, value="Fotos cadastradas")
header_fill = PatternFill(start_color='90D5AC', end_color='90D5AC', fill_type='solid')
for cell in ws_resumo['A1:B1'][0]:
    cell.fill = header_fill
for i, column in enumerate(columns_of_interest):
    ws_resumo.cell(row=i+2, column=1, value=column)
    ws_resumo.cell(row=i+2, column=2, value=sums[column])
light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
for i, column in enumerate(columns_of_interest):
    if column == 'Geral':
        for col in range(1, 3):
            cell = ws_resumo.cell(row=i+2, column=col)
            cell.fill = light_yellow_fill

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
for row in range(1, len(columns_of_interest) + 2):
    for col in range(1, 3):
        cell = ws_resumo.cell(row=row, column=col)
        cell.border = thin_border
# Seção 2: Informações no Sistema
ws_resumo.merge_cells('D1:E1')
ws_resumo.cell(row=1, column=4, value="Informações no Sistema")
header_fill = PatternFill(start_color='90D5AC', end_color='90D5AC', fill_type='solid')
for cell in ws_resumo['D1:E1'][0]:
    cell.fill = header_fill
for i, column in enumerate(columns_of_interest2):
    ws_resumo.cell(row=i+2, column=4, value=column)
    ws_resumo.cell(row=i+2, column=5, value=sums2[column])
light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
for i, column in enumerate(columns_of_interest2):
    if column == 'Cadastradas_Geral':
        for col in range(4, 6):
            cell = ws_resumo.cell(row=i+2, column=col)
            cell.fill = light_yellow_fill
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
for row in range(1, len(columns_of_interest2) + 2):
    for col in range(4, 6):
        cell = ws_resumo.cell(row=row, column=col)
        cell.border = thin_border

ws_resumo.cell(row=10, column=1).value = '=IF(E7>B6, "Há " & E7 - B6 & " placas a mais cadastradas no sistema do que fotos incluídas", "Há " & B6 - E7 & " placas a menos cadastradas no sistema do que fotos incluídas")'
wb.save(excel_name)
